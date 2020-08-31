import os
import xlsxwriter as xw
from openpyxl.utils import coordinate_to_tuple
import main_menu
import collections
from lxml import etree
import re

HOW_MANY_WORDS_IN_COUNTER = 100
MENU_RESTRICTION = 200
ROW_RESTRICTION = 3000
COL_MAX = 3000
NO_FILE = '\n' + r'没有找到文件，请确认表格在正确的路径下，并确定表格文件名称格式正确'
STANDARD_MESSAGE = 'The photo was taken in natural light because there is a slight chromatic aberration between the ' \
                   'device and the monitor, please understand'
BRAND_TO_REPLACE_KW = open('kw_brand.txt', 'r', encoding='utf-8').read()
KW_FILTER_CHAR = ': * ( ) [ ] { } ,'


def intro():
    print("\n图沓的工具\n主菜单")
    print("请选择需要的操作：")
    print('')


def cjk_detect(texts: str):
    """检测是否有中日韩文字
    :Requirement:
        re module
    :Argument:
        texts
    :Return:
        str, None
    """
    # korean
    if re.search("[\uac00-\ud7a3]", texts):
        return "ko"
    # japanese
    if re.search("[\u3040-\u30ff]", texts):
        return "ja"
    # chinese
    if re.search("[\u4e00-\u9FFF]", texts):
        return "zh"
    return None


def get_coordinate(which_cell) -> tuple:
    return coordinate_to_tuple(str(which_cell).split('.')[-1][:-1])


def validate_product() -> str:
    print('')
    print(f'当前产品类别：' + open(os.curdir + '\\product_type.txt', 'r', encoding='utf-8').read())
    wanna_change = str(input("(0：继续，1：修改)："))
    if wanna_change == str(1):
        with open(os.curdir + '\\product_type.txt', 'w', encoding='utf-8') as pt:
            pt.write(input("请输入产品类别："))
        return open(os.curdir + '\\product_type.txt', 'r', encoding='utf-8').read()
    elif wanna_change == str(0):
        return open(os.curdir + '\\product_type.txt', 'r', encoding='utf-8').read()


def get_column_except_none_cell(sheet, row_start: int, column_const: int) -> list:
    cell_list = []
    for i in range(row_start + 1, ROW_RESTRICTION):
        if sheet.cell(i, column_const).value is not None:
            sheet.cell(i, column_const).value = str(sheet.cell(i, column_const).value). \
                replace('\xa0', ' '). \
                replace('||', ' ')
            cell_list.append(sheet.cell(i, column_const))
    return cell_list


def _process_title(item_name: str) -> str:
    item_name = item_name[:200]
    if cjk_detect(item_name) is not None:
        return item_name
    new_item_name = item_name.split(' ')[:-1]
    new_item_name = ['' if each_word == ','
                     or each_word == '/'
                     or each_word == '\\'
                     else each_word
                     for each_word in new_item_name]
    new_item_name = ['' if ('(' in each_word) and (')' not in each_word)
                     or (')' in each_word) and ('(' not in each_word)
                     else each_word for each_word in new_item_name]
    new_item_name = ['' if ('[' in each_word) and (']' not in each_word)
                     or (']' in each_word) and ('[' not in each_word)
                     else each_word for each_word in new_item_name]
    new_item_name = [str(each_word).capitalize() for each_word in new_item_name]
    new_item_name = ' '.join(new_item_name).replace('  ', ' ')
    if new_item_name.endswith(',') \
            or new_item_name.endswith('*') \
            or new_item_name.endswith(' ') \
            or new_item_name.endswith(':') \
            or new_item_name.endswith('.'):
        new_item_name = new_item_name[:-1]
    return new_item_name


def process_info(sheet, info_coordinate: tuple, info, **kwargs):
    """将传入的坐标表格的值全部修改为传入的INFO"""
    if info == str(-1):
        pass
    else:
        column_offset = kwargs.get('column_offset', None)
        info_list = get_column_except_none_cell(sheet, info_coordinate[0], info_coordinate[1])
        if column_offset is not None:
            info_list = [sheet.cell(get_coordinate(each_cell)[0],
                                    get_coordinate(each_cell)[1]+column_offset) for each_cell in info_list]
        for item in info_list:
            item.value = info


def high_frequent_words(key_words_list: list):
    key_words_list = ' '.join(key_words_list)
    for each_char in KW_FILTER_CHAR.split(' '):
        key_words_list = key_words_list.replace(each_char, ' ')
    key_words_list = key_words_list.split(' ')
    key_words_list = [each_word.lower() for each_word in key_words_list]
    key_words_list = list(dict.fromkeys(key_words_list))
    kw_list = []
    for word, count in collections.Counter(key_words_list).most_common(HOW_MANY_WORDS_IN_COUNTER):
        kw_list.append(word.capitalize())
    for each_word in BRAND_TO_REPLACE_KW.split(' '):
        if each_word.capitalize() in kw_list:
            kw_list.remove(each_word.capitalize())
    return kw_list


def process_description(sheet, desc_coordinate: tuple):
    info_list = get_column_except_none_cell(sheet, desc_coordinate[0], desc_coordinate[1])
    for index, item in enumerate(info_list):
        if len(item.value) > 1500:
            info_list[index].value = ' '.join(str(item.value)[:-(len(item.value) - 1499)].split(' ')[:-1])


def cap_title(title: str, brand: str) -> str:
    title = title.split(' ')
    title = [word.capitalize() for word in title]
    title = ' '.join(title)
    title = title.replace(brand.capitalize(), brand)
    return title


def process_item_name(item_name: str) -> str:
    item_name = _process_title(item_name)
    return item_name


def process_bulletpoints(sheet, bullet_point_coordinate: tuple):
    bullet_points_dict = {'first_column': get_column_except_none_cell(sheet, bullet_point_coordinate[0],
                                                                      bullet_point_coordinate[1] + 0),
                          'second_column': get_column_except_none_cell(sheet, bullet_point_coordinate[0],
                                                                       bullet_point_coordinate[1] + 1),
                          'third_column': get_column_except_none_cell(sheet, bullet_point_coordinate[0],
                                                                      bullet_point_coordinate[1] + 2),
                          'fourth_column': get_column_except_none_cell(sheet, bullet_point_coordinate[0],
                                                                       bullet_point_coordinate[1] + 3),
                          'fifth_column': get_column_except_none_cell(sheet, bullet_point_coordinate[0],
                                                                      bullet_point_coordinate[1] + 4)
                          }
    for each_column in bullet_points_dict.keys():
        for index, each_line in enumerate(bullet_points_dict[each_column]):
            if len(each_line.value) > 500:
                bullet_points_dict[each_column][index].value = each_line.value[:499].replace('<', '(').replace('>', ')')
            if len(each_line.value) < 20:
                bullet_points_dict[each_column][index].value = STANDARD_MESSAGE


def process_price(sheet, coordinate: tuple, lowest_price: int, current_file: str):
    def update_price(input_price: int, price_loop_index: int) -> int:
        import excr_node
        excr_node.excr_node = \
            {each_key: each_value for each_key, each_value in excr_node.excr_node.items() if each_key != ''}
        for key, value in excr_node.excr_node.items():
            if key in current_file:
                if price_loop_index == 0:
                    print(f'当前文件 ({current_file})使用的汇率为：{float(value.split(",")[0].replace("!![", "").strip())}')
                price_loop_index += 1
                return int(input_price*float(value.split(',')[0].replace('!![', '').strip()))

    price = get_column_except_none_cell(sheet, coordinate[0], coordinate[1])
    price = [x for x in price if x.value != '']

    _price_loop_index = 0
    for item in price:
        updated_price = update_price(lowest_price, _price_loop_index)
        _price_loop_index += 1
        if int(float(item.value)) < 1:
            # 删除低于最低价格的行
            print(f"表格中{item}的价格过低于1，正在删除")
            for col in range(1, COL_MAX):
                sheet.cell(coordinate_to_tuple(str(item).split('.')[-1][:-1])[0], col).value = None
        elif (updated_price is not None) and (int(float(item.value)) < updated_price):
            print(f"文件({current_file})的表格中{item}的价格为{item.value}低于{updated_price}，"
                  f"更改为{updated_price}，"
                  f"如果需要别的数值请在表格中自行更改")
            item.value = str(int(updated_price))
        else:
            item.value = str(int(int(str(item.value).split('.')[0])) - 1)


def make_menu(functions: dict) -> None:
    r"""创建菜单函数，将会执行选中的函数

    传入dict，key为描述，value为需要执行的函数
    **将会执行选中的函数

    :Return:
        None
    """
    _menu = {}
    index = 0
    os.system('cls')
    for descreption, func in functions.items():
        _menu[(index, descreption)] = func
        index = index + 1

    for item in _menu.keys():
        print(str(item[0]) + '\t' + item[1])
        print('')

    ui = input('输入选项：')
    _menu_options = [str(_index_desc[0]) for _index_desc in _menu.keys()]
    if ui not in _menu_options:
        input(f'没有找到输入的选项: {ui} (回车继续)')
        make_menu(functions)

    if len(ui.split(' ')) > 1:
        for each_ui in ui.split(' '):
            for index_desc, func in _menu.items():
                if each_ui == str(index_desc[0]):
                    func()
    else:
        for index_desc, func in _menu.items():
            if ui == str(index_desc[0]):
                func()


def make_menu_part_functions(functions: dict):
    r"""创建菜单函数，仅返回函数名称，不执行函数

    传入functions: dict
    key为描述，value函数的名称

    :Return:
        func_name  函数的名称
    """
    _menu = {}
    index = 0
    for descreption, func_name in functions.items():
        _menu[(index, descreption)] = func_name
        index = index + 1

    for item in _menu.keys():
        print(str(item[0]) + '\t' + item[1])
    ui = input('输入选项：')
    if len(ui.split(' ')) > 1:
        for each_ui in ui.split(' '):
            for item, func_name in _menu.items():
                if each_ui == str(item[0]):
                    return func_name
    else:
        for item, func_name in _menu.items():
            if ui == str(item[0]):
                return func_name


# 索引用户输入的文件夹中的文件
def index_files(**kwargs) -> tuple:
    r"""请求用户输入一个文件夹，返回文件夹的路径和选择的文件路径

    :Keyword Arguments:
        ui_msg: 请求输入是提示的字符串
    :Return:
        (文件夹路径，文件路径)
    """
    folder = input(kwargs.get('ui_msg', '输入包含表格的文件夹：'))
    files = {}
    index = 0
    if folder == '-1':
        back_to_main_menu()
    elif not os.path.isdir(folder):
        print('请确保输入的是一个文件夹路径')
        index_files()
    for folder, subfolder, file in os.walk(folder):
        for each_file in file:
            files[index] = folder + '\\' + each_file
            index += 1
    print("当前文件夹中包含的文件有：")
    for index, file in files.items():
        print(index, end='')
        print('\t' + file.split('\\')[-1])
    ui = str(input(kwargs.get('which_file_msg', '请选择文件：'))).strip()
    which_file = ''
    for selection in files.keys():
        if len(ui.split(' ')) > 1:
            which_file = []
            for each_ui in ui.split(' '):
                which_file.append(files[int(each_ui)])
        if ui == str(selection):
            which_file = files[selection]
    return folder, which_file


def back_to_main_menu(**kwargs):
    r"""返回主菜单
    :Keyword Arguments:
            enter_quit: 按回车返回
    """
    if kwargs.get('enter_quit', False):
        input('\n输入回车返回主菜单')
    os.system('cls')
    main_menu.main_menu()
    return 0


# 使用统一参数处理多个文件
def multiple_file_process(process_class, class_parameter: dict, **pas_args):
    """可以传入的选项：method_para, process_method, 自动调用save_sheet """
    folder, which_file = index_files()
    if type(which_file) is list:
        for each_file in which_file:
            print('\n开始处理表格：' + each_file.split("\\")[-1])
            pas = process_class(each_file, class_parameter)
            pas_args['method_para'] = pas_args.get('method_para', None)
            if pas_args['method_para'] is None:
                pas.__getattribute__(pas_args['process_method'])()
            else:
                pas.__getattribute__(pas_args['process_method'])(*pas_args['method_para'])
            pas.__getattribute__('save_sheet')(folder, each_file.split('\\')[-1])
    else:
        print('开始处理表格：' + which_file.split("\\")[-1])
        pas = process_class(which_file, class_parameter)
        pas_args['method_para'] = pas_args.get('method_para', None)
        if pas_args['method_para'] is None:
            pas.__getattribute__(pas_args['process_method'])()
        else:
            pas.__getattribute__(pas_args['process_method'])(*pas_args['method_para'])
        pas.__getattribute__('save_sheet')(folder, which_file.split('\\')[-1])
    input("处理完成")


def main_menu_quit():
    try:
        os.system('exit')
    except Exception as e:
        print(e)

        class QuitMainMenu(Exception):
            pass
        raise QuitMainMenu('退出程序')


def asin_price_menu():
    _menu = {'回主菜单': back_to_main_menu,
             '爬取ASIN，价格和主图链接': get_asin_price}
    make_menu(_menu)


def get_asin_price():
    folder, which_file = index_files()
    out_path = '\\'.join(str(which_file).split('\\')[:-1])+'\\输出文件_'+str(which_file).split("\\")[-1]+'.xlsx'
    listing_xpath = '//*[@id="search"]/div[1]/div[2]/div/span[3]/div[2]'
    title = '//h2/a/span/text()'
    ads = 'aok-inline-block s-sponsored-label-info-icon'
    html = etree.HTML(open(which_file, 'r', encoding='utf-8').read(), etree.HTMLParser())
    all_items = html.xpath(listing_xpath)[0][:-3]
    print(len(all_items))
    for index, each_listing in enumerate(all_items):
        if ads in str(etree.tostring(each_listing)):
            all_items[index] = None
    all_items = [item for item in all_items if item is not None]
    print(len(all_items))
    listing = {}
    for each_listing in all_items:
        each_listing = etree.HTML(etree.tostring(each_listing), etree.HTMLParser())
        listing[','.join(each_listing.xpath('//@data-asin'))] = \
            {'asin': ','.join(each_listing.xpath('//@data-asin')),
             'title': ','.join(each_listing.xpath(title)),
             'image': ','.join(each_listing.xpath('//img/@src')).split('AC')[0]+'_SL1024_.jpg',
             'price': ','.join(each_listing.xpath('//span[@class="a-offscreen"]/text()')).replace('\xa0', ' ')
             }
    out_file = xw.Workbook(out_path)
    out_sheet = out_file.add_worksheet('sheet1')
    row = 0
    col = 0
    for k, v in listing.items():
        for i, j in v.items():
            print(j)
            out_sheet.write(row, col, j)
            col += 1
            if col == len(v):
                col = 0
        row += 1
    out_file.close()
    back_to_main_menu()


if __name__ == '__main__':
    while True:
        get_asin_price()
    pass
