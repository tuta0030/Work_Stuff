import openpyxl


COLUMN_RANGE_RESTRICTION = 2000
ROW_RANGE_RESTRICTION = 10


class LoadAmazonSheet(object):

    def __init__(self, sheet_path: str):
        self.wb = openpyxl.load_workbook(sheet_path)
        self.sheet = self.wb['sheet1']
        self.current_file = sheet_path

        def find_cell(sheet, cell_name: str):
            for _r in range(1, ROW_RANGE_RESTRICTION):
                for _c in range(1, COLUMN_RANGE_RESTRICTION):
                    if sheet.cell(_r, _c).value == cell_name:
                        return sheet.cell(_r, _c)

        # find "item_name" cell
        self.item_name_cell = find_cell(self.sheet, 'item_name')

        # find "bullet_point1" cell
        self.bullet_point_cell = find_cell(self.sheet, 'bullet_point1')

        # find "standard_price" cell
        self.price_cell = find_cell(self.sheet, 'standard_price')

        # find "node" cell
        self.node_cell = find_cell(self.sheet, 'recommended_browse_nodes')

        # find "keywords" cell
        self.keywords_cell = find_cell(self.sheet, 'generic_keywords1')

        # find "item_type" cell
        self.item_type_cell = find_cell(self.sheet, 'item_type')

        # find "product_description" cell
        self.description = find_cell(self.sheet, 'product_description')

        # find "update_delete" cell
        self.update_delete = find_cell(self.sheet, 'update_delete')

        # find "color_name" cell
        self.color_name = find_cell(self.sheet, 'color_name')

        # find "size_name" cell
        self.size_name = find_cell(self.sheet, 'size_name')

        # find "fulfillment_latency" cell
        self.fulfillment_latency = find_cell(self.sheet, 'fulfillment_latency')

        # find "part_number" cell
        self.part_number_cell = find_cell(self.sheet, 'manufacturer')

        # find "model" cell
        self.model_cell = find_cell(self.sheet, 'manufacturer')


if __name__ == '__main__':
    pass
