# -*- coding: utf-8 -*-
# @Time    : 2020/7/28 13:42
# @Author  : Eric Shen
# @Email   : Eric_Shenarrzine@outlook.com
# @File    : text_process.py
# @Software: PyCharm

import pas_utility as pasu
import os
import re
import htm_file_warp
import openpyxl
import datetime
import traceback

ROW_RANGE_RESTRICTION = 2000
COLUMN_RANGE_RESTRICTION = 2000
BR_PATTERN = '$$$'
SEPARATOR = '^^^'
EXCHANGE_RATE_NODE = ('!![', ']!!')
EXCR_NODE_TEMP = {'ES': '', 'JP': '', 'IT': '', 'US': '', 'FR': '',
                  'DE': '', 'NL': '', 'SA': '', 'AE': '', 'CA': '', 'MX': ''}


class Translate:

    def __init__(self):
        self.file_directory = ''
        self.sheet = ''

    def load_sheet(self, which_file) -> dict:
        ws = openpyxl.load_workbook(which_file).get_sheet_by_name('sheet1')
        self.sheet = ws
        item_name_cell = find_cell(ws, 'item_name')
        bullet_point_cell = find_cell(ws, 'bullet_point1')
        color_name = find_cell(ws, 'color_name')
        color_map = find_cell(ws, 'color_map')
        size_name = find_cell(ws, 'size_name')
        size_map = find_cell(ws, 'size_map')
        description = find_cell(ws, 'product_description')
        sheet_cells = {'标题': item_name_cell,
                       '五点': bullet_point_cell,
                       '描述': description,
                       '变体-颜色': color_name,
                       '变体-尺寸': size_name,
                       '变体-颜色map': color_map,
                       '变体-尺寸map': size_map
                       }
        return sheet_cells

    @staticmethod
    def add_cor(each_cell) -> str:
        """add coordinate for cell content"""
        return f'{str(pasu.get_coordinate(each_cell))} {SEPARATOR} ' + str(each_cell.value)

    def get_all_column(self, sheet, which_content):
        """which_content should be the cell"""
        content_coordinate = pasu.get_coordinate(which_content)
        if which_content.value == 'bullet_point1':
            bullet_points_dict = {'first_column': pasu.get_column_until_none_cell(sheet, content_coordinate[0],
                                                                                  content_coordinate[1] + 0),
                                  'second_column': pasu.get_column_until_none_cell(sheet, content_coordinate[0],
                                                                                   content_coordinate[1] + 1),
                                  'third_column': pasu.get_column_until_none_cell(sheet, content_coordinate[0],
                                                                                  content_coordinate[1] + 2),
                                  'fourth_column': pasu.get_column_until_none_cell(sheet, content_coordinate[0],
                                                                                   content_coordinate[1] + 3),
                                  'fifth_column': pasu.get_column_until_none_cell(sheet, content_coordinate[0],
                                                                                  content_coordinate[1] + 4)
                                  }
            full_bp_list = []
            for key, value in bullet_points_dict.items():
                content_list = \
                    [self.add_cor(each_cell) for each_cell in value]
                full_bp_list.append(self.htm_warp(content_list))
            return '\n'.join(full_bp_list)
        content_list = pasu.get_column_until_none_cell(sheet,
                                                       content_coordinate[0],
                                                       content_coordinate[1])
        content_list = \
            [self.add_cor(each_cell).replace('<br>', BR_PATTERN)
             for each_cell in content_list]
        return self.htm_warp(content_list)

    def save_as_html(self, what, content: str):
        out_path = '\\'.join(self.file_directory.split('\\')[:-1])
        with open(out_path + '\\' + what + '.htm', 'a', encoding='utf-8') as file:
            file.write(htm_file_warp.htm_file_head)
            file.write('\n')
            file.write(content)
            file.write('\n')
            file.write(htm_file_warp.htm_file_tail)
        print(f'保存 {what + ".htm"} 到 {self.file_directory}')

    @staticmethod
    def htm_warp(content: list) -> str:
        def add_htm_warp(title):
            return f"""<tr height="18" style='height:13.50pt;'> <td class="xl65" height="18" colspan="14" 
        style='height:13.50pt;mso-ignore:colspan;' x:str>{title}</td> <td></td> </tr> """

        all_titles = [add_htm_warp(each_title) for each_title in content]
        return '\n'.join(all_titles)

    def save_all(self):
        self.file_directory = pasu.index_files()[-1]
        if type(self.file_directory) is list:
            input('只能输入一个文件路径，请确保只选择了一个文件（回车返回主菜单）')
            pasu.back_to_main_menu()
        content_dict = self.load_sheet(self.file_directory)
        for key, value in content_dict.items():
            save_this_content = self.get_all_column(self.sheet, value)
            self.save_as_html(f'所有内容_{datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S")}', save_this_content)
        pasu.back_to_main_menu()


# 从文本文件生成表格文件
class ReadTranslatedTxt(object):

    def __init__(self):
        self.directory = ''
        self.langs = []
        self.langs_dict = {}

    @staticmethod
    def calculate_time_exchange_rate(each_price, exchange_rate):
        r""" each_price:cell, exchange_rate:int, float, str"""
        # print(f'价格为 {float(each_price.value)}\t 汇率为: {float(exchange_rate)} '
        #       f'\n结果为 {int(float(each_price.value) * float(exchange_rate))}')
        return int(float(each_price.value) * float(exchange_rate))

    def find_all_txt_file(self) -> list:
        self.directory = input('输入 <翻译好的txt文件> 所在的路径:')
        files = []
        if self.directory == '-1':
            pasu.back_to_main_menu()
        elif not os.path.isdir(self.directory):
            print('请输入一个文件夹路径')
            self.find_all_txt_file()
        for folder, subfolder, file in os.walk(self.directory):
            for each_file in file:
                if each_file.split('.')[-1] == 'txt':
                    files.append(folder + '\\' + each_file)
        return files

    def get_langs(self, files: list):
        langs = []
        [langs.append(str(each_htm).replace('五点', '').replace('标题', '').replace('描述', '')
                      .split('\\')[-1].replace('.txt', ''))
         for each_htm in files]
        langs = list(set(langs))
        self.langs = langs

    def get_langs_dict(self, files: list):
        for each_lang in self.langs:
            self.langs_dict[each_lang] = []
            for each_file in files:
                if each_lang in each_file:
                    self.langs_dict[each_lang].append(each_file)

    def get_langs_and_langs_dict(self, files):
        self.get_langs(files)
        self.get_langs_dict(files)

    def main(self):
        files = self.find_all_txt_file()
        self.get_langs_and_langs_dict(files)
        oc_file = pasu.index_files(ui_msg='输入 <表格文件> 所在的路径:', which_file_msg='请选择原始表格文件:')[-1]
        original_wb = openpyxl.load_workbook(str(oc_file))
        original_sheet = original_wb.get_sheet_by_name('sheet1')

        # 获取价格和节点的成员函数
        def get_node_price_list():
            _new_wb = openpyxl.load_workbook(str(oc_file))
            _new_sheet = _new_wb.get_sheet_by_name('sheet1')
            _node_list = get_content_list(_new_sheet, 'recommended_browse_nodes')
            _price_list = get_content_list(_new_sheet, 'standard_price')
            _price_list = [each_cell for each_cell in _price_list if each_cell.value != '']
            return _node_list, _price_list, _new_wb, _new_sheet

        if check_if_the_same_day(get_time_stamp()):
            excr_node = {}
            excr_node_result = asking_for_excr_node_input(EXCR_NODE_TEMP)
            excr_node_result = excr_node_result.split('\n')
            excr_node_result = [each_line for each_line in excr_node_result if each_line != '' or each_line != ':']
            for each_line in excr_node_result:
                excr_node[each_line.split(':')[0].strip()] = \
                    EXCHANGE_RATE_NODE[0]+str(each_line.split(':')[-1].strip())+EXCHANGE_RATE_NODE[1]
            with open('excr_node.py', 'w', encoding='utf-8') as t:
                t.write('excr_node = '+str(excr_node))
            with open('_time_stamp_for_excr_node.py', 'w', encoding='utf-8') as t:
                t.write('time_stamp = ' +
                        f'"{datetime.datetime.strftime(datetime.datetime.now(), "%Y, %m, %d, %I, %M, %S")}"')

        # 处理文本文件
        for lang, file_list in self.langs_dict.items():
            import excr_node
            template = {key: value for key, value in excr_node.excr_node.items() if key != ''}
            for each_file in file_list:
                if excr_node.excr_node is not None:
                    for lang_excr_node, _excr_node in template.items():
                        if lang_excr_node in each_file.split('\\')[-1]:
                            line_prepender(each_file, _excr_node)

                content = open(each_file, encoding='utf-8').read()
                content_list = content.split('\n')
                content_list = [each_line for each_line in content_list if each_line != '']
                content_list = [each_line for each_line in content_list if SEPARATOR in each_line]
                for each_content in content_list:
                    try:
                        if len(each_content.split(SEPARATOR)[0]) > 9:
                            continue
                        each_content = str(each_content).split(SEPARATOR)
                        row = int(each_content[0].strip()[1:-1].replace('、', ',').split(',')[0])
                        col = int(each_content[0].strip()[1:-1].replace('、', ',').split(',')[1])
                        original_sheet.cell(row, col).value = each_content[-1].strip().replace(BR_PATTERN, ' <br> ')\
                            .replace('$$ $', ' <br> ').replace('$ $$', ' <br> ')
                    except Exception as e:
                        print(f'{each_file} 中的内容： {each_content} 发生了错误 {e}')
                        continue

                if EXCHANGE_RATE_NODE[0] not in content:
                    input(f'\n文本文件: ({each_file}) 当中没有标明汇率和节点，请检查文件（回车继续）')
                    continue
                elif EXCHANGE_RATE_NODE[0] in content:
                    search_result = re.search(re.compile(r'(?<=!!\[).+(?=]!!)'), content)
                    # print(f'\n{each_file} 找到的!![]!! {search_result[0]}')
                    if search_result is None:
                        continue
                    exchange_rate, node = str(re.search(re.compile(r'(?<=!!\[).+(?=]!!)'), content)[0]).split(',')
                    node_list, price_list, new_wb, new_sheet = get_node_price_list()
                    for each_node in node_list:
                        row, col = pasu.get_coordinate(each_node)
                        original_sheet.cell(int(row), int(col)).value = str(node).strip()
                    for each_price in price_list:
                        row, col = pasu.get_coordinate(each_price)
                        original_sheet.cell(int(row), int(col)).value = \
                            self.calculate_time_exchange_rate(each_price, exchange_rate)
                    print(f'\n当前的语言: {lang}')
                    print(f'当前使用的节点：{node}')
                    print(f'当前使用的汇率:{exchange_rate}')
                else:
                    class NoExchangeNodeError(Exception):
                        pass
                    raise NoExchangeNodeError('No exchange rate and node')

                out_file_name = self.directory + '\\' + lang + '_' + str(oc_file).split('\\')[-1]
                print(f'正在处理  {out_file_name}')
                original_wb.save(out_file_name)

        pasu.back_to_main_menu(enter_quit=True)


# 全局函数
def get_content_list(sheet, cell_name: str) -> list:
    cell = find_cell(sheet, cell_name)
    cell_coordinate = pasu.get_coordinate(cell)
    content_list = pasu.get_column_until_none_cell(sheet, cell_coordinate[0], cell_coordinate[1])
    return content_list


def find_cell(sheet, cell_name: str):
    for _r in range(1, ROW_RANGE_RESTRICTION):
        for _c in range(1, COLUMN_RANGE_RESTRICTION):
            if sheet.cell(_r, _c).value == cell_name:
                return sheet.cell(_r, _c)


def check_if_the_same_day(_time_stamp: datetime.datetime):
    if _time_stamp.day == datetime.datetime.now().day and \
       _time_stamp.year == datetime.datetime.now().year and \
       _time_stamp.month == datetime.datetime.now().month:
        return False
    elif (datetime.datetime.now().day - _time_stamp.day) > 1:
        return True
    else:
        return True


def line_prepender(filename, line):
    with open(filename, 'r+', encoding='utf-8', errors='ignore') as f:
        content = f.read()
        if line not in content:
            if '!![' in content:
                content = re.sub(re.compile(r'!!\[.*]!!'), ' ', content)
            f.seek(0, 0)
            f.writelines(line.rstrip('\r\n') + '\n' + content)


def get_time_stamp() -> datetime.datetime:
    if os.path.isfile('_time_stamp_for_excr_node.py'):
        import _time_stamp_for_excr_node
        try:
            _time_stamp = _time_stamp_for_excr_node.time_stamp.split(',')
            _time_stamp = [int(each_part) for each_part in _time_stamp]
            return datetime.datetime(*_time_stamp)
        except Exception as e:
            print(f'\n读取时间戳文件失败 {e}')
            return datetime.datetime(2000, 1, 1)
    else:
        return datetime.datetime(2000, 1, 1)


def asking_for_excr_node_input(_excr_node_temp: dict) -> str:
    with open('ExchangeRate_Node.txt', 'w', encoding='utf-8') as f:
        for key, value in _excr_node_temp.items():
            f.write(f'{key}:{value}\n')
    os.startfile('ExchangeRate_Node.txt')
    is_finished = input('是否完成输入(Y/N):')
    if (is_finished == 'y') and (open('ExchangeRate_Node.txt', 'r', encoding='utf-8').read() is not None):
        return open('ExchangeRate_Node.txt', 'r', encoding='utf-8').read()
    else:
        print('未完成输入，尝试重新输入...')
        file = open('ExchangeRate_Node.txt', 'r', encoding='utf-8')
        file.close()
        asking_for_excr_node_input(_excr_node_temp)


# 文本处理模块主函数
def main():
    translate = Translate()
    readtranslate = ReadTranslatedTxt()
    _menu = {'返回主菜单': pasu.back_to_main_menu,
             '通过表格保存htm文件': translate.save_all,
             '通过txt生成新的表格文件': readtranslate.main}
    pasu.make_menu(_menu)


if __name__ == '__main__':
    check_if_the_same_day(get_time_stamp())
