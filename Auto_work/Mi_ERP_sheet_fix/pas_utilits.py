import os
import datetime
import re
from openpyxl.utils import coordinate_to_tuple

ROW_RESTRICTION = 2000
COL_MAX = 2000
MAIN_PATH = open(os.curdir + '\\path.txt', 'r', encoding='utf-8').read()
PRODUCT_TYPE = open(os.curdir + '\\product_type.txt', 'r', encoding='utf-8').read()
INTRO = r"""
图沓的亚马逊表格处理工具
请注意文件命名方式：
    1. <文件夹> 命名方式: <日期_产品名称>  例如（根目录\20200601_游泳圈）
    2. <文件> 命名方式: <产品名称+国家_亚马逊表_日期>  例如（游泳圈UK_亚马逊表_20200601.xlsx）
    3. 分类节点和关键词选项如果已经填写，可以按-1跳过
""" + f'当前设置的主路径为：{MAIN_PATH}\n'
NO_FILE = '\n' + r'没有找到文件，请确认表格在正确的路径下，并确定表格文件名称格式正确'
STANDARD_MESSAGE = 'The photo was taken in natural light because there is a slight chromatic aberration between the ' \
                   'device and the monitor, please understand '


def get_date() -> str:
    time = ''.join(str(datetime.datetime.now()).split(' ')[0].split('-'))
    return str(time)


def validate_main_path() -> str:
    if os.path.isdir(MAIN_PATH) is not True:
        main_path = input("查找主路径失败，请输入包含产品文件夹的路径：")
        with open(os.curdir + '\\path.txt', 'w', encoding='utf-8') as p:
            p.write(main_path)
        _main_path = open(os.curdir + '\\path.txt', 'r', encoding='utf-8').read()
        return _main_path
    else:
        return MAIN_PATH


def validate_product() -> str:
    print('')
    print(f'当前产品类别：{PRODUCT_TYPE}')
    wanna_change = str(input("(0：继续，1：修改)："))
    if wanna_change == str(1):
        with open(os.curdir + '\\product_type.txt', 'w', encoding='utf-8') as pt:
            pt.write(input("请输入产品类别："))
        return open(os.curdir + '\\product_type.txt', 'r', encoding='utf-8').read()
    elif wanna_change == str(0):
        return PRODUCT_TYPE


def select_time() -> str:
    which_time = int(input("选择<需要处理的文件>名中的时间：0:默认时间（当前日期）， 1:自定义时间："))
    if which_time == 0:
        time = get_date()
        return time
    elif which_time == 1:
        time = str(input("请输入需要设置的时间(格式:年月日-比如：20200501 可省略年份)："))
        if len(time) == 4:
            time = str(datetime.datetime.now())[:4] + time
        return time
    else:
        print("只能输入0和1")
        select_time()


def get_column_until_none_cell(sheet, row_start: int, column_const: int) -> list:
    cell_list = []
    for i in range(row_start + 1, ROW_RESTRICTION):
        if sheet.cell(i, column_const).value is not None:
            sheet.cell(i, column_const).value = str(sheet.cell(i, column_const).value). \
                replace('\xa0', ' '). \
                replace('||', ' ')
            cell_list.append(sheet.cell(i, column_const))
    return cell_list


def __random_title(item_name: str) -> str:
    _pattern = re.compile(r'\s°')
    length = len(item_name)
    item_name = re.sub(_pattern, '°', item_name)  # 移除°前的空格

    item_name_h = item_name.split(' ')[:-round(len(item_name.split(' ')) / 2)]
    item_name_t = item_name.split(' ')[round(len(item_name.split(' ')) / 2):]

    color_size = []

    # 将带有","的变体数值保留
    for word in item_name_t:
        if ',' in word or '(' in word or ')' in word:
            try:
                color_size.append(' '.join(item_name_t[item_name_t.index(word):]))
                item_name_t = item_name_t[:item_name_t.index(word)]
                if length > 199:
                    item_name_t = []
            except ValueError:
                item_name_t = []

    rand_item_name_t = list(set(item_name_t))
    rand_item_name_t = rand_item_name_t + color_size
    new_item_name = item_name_h + rand_item_name_t
    new_item_name = ' '.join(new_item_name).replace('  ', ' ')
    return new_item_name


def process_info(sheet, info_coordinate: tuple, info):
    """将传入的坐标表格的值全部修改为传入的INFO"""
    if info == str(-1):
        pass
    else:
        info_list = get_column_until_none_cell(sheet, info_coordinate[0], info_coordinate[1])
        for index, item in enumerate(info_list):
            info_list[index].value = info


def process_description(sheet, desc_coordinate: tuple):
    info_list = get_column_until_none_cell(sheet, desc_coordinate[0], desc_coordinate[1])
    for index, item in enumerate(info_list):
        if len(item.value) > 1500:
            info_list[index].value = ' '.join(str(item.value)[:-(len(item.value) - 1499)].split(' ')[:-1])


def process_item_name(item_name: str) -> str:
    # random words
    item_name = __random_title(item_name)

    return item_name


def process_bulletpoints(sheet, bullet_point_coordinate: tuple):
    bullet_points_dict = {'first_column': get_column_until_none_cell(sheet, bullet_point_coordinate[0],
                                                                     bullet_point_coordinate[1] + 0),
                          'second_column': get_column_until_none_cell(sheet, bullet_point_coordinate[0],
                                                                      bullet_point_coordinate[1] + 1),
                          'third_column': get_column_until_none_cell(sheet, bullet_point_coordinate[0],
                                                                     bullet_point_coordinate[1] + 2),
                          'fourth_column': get_column_until_none_cell(sheet, bullet_point_coordinate[0],
                                                                      bullet_point_coordinate[1] + 3),
                          'fifth_column': get_column_until_none_cell(sheet, bullet_point_coordinate[0],
                                                                     bullet_point_coordinate[1] + 4)
                          }
    for index, each_line in enumerate(bullet_points_dict['first_column']):
        if len(each_line.value) > 500:
            bullet_points_dict['first_column'][index] = each_line[:499].replace('<', '(').replace('>', ')')
    for index, each_line in enumerate(bullet_points_dict['second_column']):
        if len(each_line.value) > 500:
            bullet_points_dict['first_column'][index] = each_line[:499].replace('<', '(').replace('>', ')')
    for index, each_line in enumerate(bullet_points_dict['third_column']):
        if len(each_line.value) > 500:
            bullet_points_dict['first_column'][index] = each_line[:499].replace('<', '(').replace('>', ')')
    for index, each_line in enumerate(bullet_points_dict['fourth_column']):
        if len(each_line.value) > 500:
            bullet_points_dict['first_column'][index] = each_line[:499].replace('<', '(').replace('>', ')')
    for index, each_line in enumerate(bullet_points_dict['fifth_column']):
        if len(each_line.value) > 500:
            bullet_points_dict['first_column'][index] = each_line[:499].replace('<', '(').replace('>', ')')


def process_price(sheet, coordinate: tuple, exchange_rate: float):
    price = get_column_until_none_cell(sheet, coordinate[0], coordinate[1])
    lowest_price = int(input("输入最低价格："))
    for index, item in enumerate(price):
        price[index].value = str(int(int(str(item.value).split('.')[0]) * exchange_rate) - 1)
        if int(price[index].value) < lowest_price:
            # 删除低于最低价格的行
            for col in range(1, COL_MAX):
                sheet.cell(coordinate_to_tuple(str(item).split('.')[-1][:-1])[0], col).value = None


if __name__ == '__main__':
    pass
