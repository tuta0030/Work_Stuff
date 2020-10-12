import pas_utility
import openpyxl
import xlsxwriter
import os

_meta_path = r'D:\上传表格文件\test\auto_all'
AIO_FILE_NAME = r'\_AIO_FILE.xlsx'
AIO_PATH = pas_utility.asking_for_dir_path(ui_msg='请输入全部表格文件需要输出的文件夹：')+AIO_FILE_NAME


def merge_all_sheet(file_list, **kwargs) -> None:
    """合并所有表格，保留前三行，合并剩下的所有行
    :Requirement:
        openpyxl, xlsxwriter
    :Arguments:
        file_list, **kwargs['ignore_error']
    :Return:
        None
    """
    def main():
        all_in_one = []

        def all_into_list(all_in_one_list, _file_list):
            _process_file_count = 0
            row_off_set = 0
            for file in _file_list:
                wb = openpyxl.load_workbook(file)
                sheet = wb['sheet1']
                print('\n', file)
                print('max row:', sheet.max_row)
                print('max column:', sheet.max_column)

                if not row_off_set:
                    start_row = 1
                else:
                    start_row = 4
                for row in range(start_row, sheet.max_row+1):
                    for col in range(1, sheet.max_column+1):
                        all_in_one_list.append(((row+row_off_set, col), sheet.cell(row, col).value))
                row_off_set += sheet.max_row-3
                _process_file_count += 1
                print('当前处理文件：', _process_file_count)
            print('共处理文件：', _process_file_count)
            return all_in_one_list

        all_in_one = all_into_list(all_in_one, file_list)
        out_wb = xlsxwriter.Workbook(AIO_PATH)
        out_ws = out_wb.add_worksheet('sheet1')
        for _ in all_in_one:
            out_ws.write(_[0][0]-1, _[0][1]-1, _[1])
        out_wb.close()
        print('处理完成')

    if kwargs['ignore_error']:
        try:
            main()
        except Exception as e:
            print(e)
    else:
        main()


def check_sub_folders(meta_path) -> tuple:
    """Check if there are sub folders.

    Pass in a path wanna to check,
    return a tuple (bool, list)
    True if there are sub folders, list of the sub folders path
    False if there are none sub folders, list of the files path

    """
    files = []
    folders = []
    for dir_path, dir_names, dir_files in os.walk(meta_path):
        for f_name in dir_files:
            files.append(os.path.join(dir_path, f_name))
        for d_name in dir_names:
            folders.append(os.path.join(dir_path, d_name))
    if not folders:
        return False, files
    else:
        return True, folders


if __name__ == '__main__':
    # print('Starting All In One Process...')
    # print(check_sub_folders(_meta_path))
    merge_all_sheet(pas_utility.index_files(which_file_msg='选择需要合并的文件：')[1], ignore_error=True)
