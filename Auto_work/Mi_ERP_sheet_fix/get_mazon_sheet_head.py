import openpyxl
import os

def load_amazon_header(file_path):
    book = openpyxl.load_workbook(file_path)
    print(book.sheetnames)
    # sheet = book[book.sheetnames[0]]
    # print(type(sheet))
    # print(sheet)



if __name__ == '__main__':
    load_amazon_header(os.curdir+'\\Amazon_header.xlsx')
