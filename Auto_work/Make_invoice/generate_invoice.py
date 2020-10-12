import openpyxl
import random
import os
from lxml import etree
from Auto_work.Make_invoice import invoice_parse_page, selector_path
from Auto_work.Make_invoice.order_info import info
from Auto_work.Mi_ERP_sheet_fix import pas_utility as pasu
import pyperclip

first_out_path = True
out_path = ''


def addToClipBoard():
    command = 'echo ' + info["order_id"] + '| clip'
    os.system(command)


class Invoice(object):

    def __init__(self):
        self.inovice_path = os.curdir+'\\发票.xlsx'
        self.PINO_num_path = os.curdir+'\\PINO_num.txt'         # Order_Page_two_order.html
        self.HTML_bs4 = os.curdir+'\\Order_Page.html'  # Order_Page_with_phone  Order_Page
        self.HTML_etree = os.curdir+'\\Order_Page.html'  # Order_Page_clothes Order_Page_with_no_buyer_info_atall
        self.manual_info_msg = '需要手动添加'
        self.book = openpyxl.load_workbook(self.inovice_path)
        self.sheet = ''
        self.xpath = selector_path.path_xpath
        self.info = info
        self.sheet_info = {}

    def get_html_etree(self):
        if '<html' not in pyperclip.paste():
            print('没有在剪贴板中找到html标签，请确保复制了整个html页面的文本')
            again()
        with open(self.HTML_etree, 'w', encoding='utf-8') as f:
            f.write(pyperclip.paste())
        src = open(self.HTML_etree, 'r', encoding='utf-8').read()
        html = etree.HTML(src, etree.HTMLParser())
        return html

    def set_PINO(self, sheet):
        self.info['PINO'] = sheet['E5'].value = str(random.randint(10000000, 99999999))
        if not os.path.isfile(self.PINO_num_path):
            with open(self.PINO_num_path, 'w', encoding='utf-8') as f:
                f.write('')
        with open(self.PINO_num_path, 'r', encoding='utf-8') as n:
            nums = n.read()
            for self.info['PINO'] in nums:
                self.info['PINO'] = sheet['E5'].value = str(random.randint(10000000, 99999999))
            else:
                with open(self.PINO_num_path, 'a', encoding='utf-8') as a:
                    a.write(self.info['PINO'])
                    a.write('\n')

    def read_invoice_content(self):
        self.book = openpyxl.load_workbook(self.inovice_path)
        self.sheet = self.book[self.book.sheetnames[0]]
        self.sheet_info['buyer_info'] = self.sheet['A4']
        self.sheet_info['date'] = self.sheet['D5']
        self.sheet_info['PINO'] = self.sheet['E5']
        self.sheet_info['SKU'] = self.sheet['A10']
        self.sheet_info['product_name'] = self.sheet['C10']
        self.sheet_info['price'] = self.sheet['D10']
        self.sheet_info['quantity'] = self.sheet['E10']
        self.sheet_info['amount'] = self.sheet['F10']
        self.sheet_info['ship_cost'] = self.sheet['F11']
        self.sheet_info['sub_total'] = self.sheet['F12']
        self.sheet_info['signature'] = self.sheet['A13']
        self.sheet_info['total'] = self.sheet['F14']
        self.sheet_info['order_id'] = info['order_id']

    def write_invoice_content(self):
        self.set_PINO(self.sheet)
        invoice_parse_page.parse_page_lxml(self, self.sheet_info, self.sheet)

    def save_by_order(self):
        addToClipBoard()
        global first_out_path
        if first_out_path:
            _out_path = input('输入输出文件目录：')
            global out_path
            out_path = _out_path
            self.book.save(_out_path+f"\\{info['order_id']}.xlsx")
            self.book.close()
            os.startfile(_out_path+f"\\{info['order_id']}.xlsx")
            first_out_path = False
        else:
            self.book.save(out_path + f"\\{info['order_id']}.xlsx")
            self.book.close()
            os.startfile(out_path + f"\\{info['order_id']}.xlsx")
        again()

    def main(self):
        self.read_invoice_content()
        self.write_invoice_content()
        self.save_by_order()


def again():
    _again = input('是否再次生成？(y/n):')
    if _again == 'y':
        main()
    elif _again == 'n':
        pasu.back_to_main_menu(enter_quit='回车返回主菜单')
    else:
        pasu.back_to_main_menu(enter_quit='无法识别的输入，返回主菜单')


def main():
    _i = Invoice()
    _i.main()
    pasu.back_to_main_menu()


if __name__ == '__main__':
    i = Invoice()
    i.main()
